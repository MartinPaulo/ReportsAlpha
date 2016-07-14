import csv
import os
import re
from datetime import date
from wsgiref.util import FileWrapper

import django
from dateutil.relativedelta import relativedelta
from django.apps import AppConfig
from django.conf import settings
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import render
from django.utils import timezone
from django.views import generic
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from reports.fake_data import factory
from .models import Report

first_cap_re = re.compile('(.)([A-Z][a-z]+)')
all_cap_re = re.compile('([a-z0-9])([A-Z])')


def convert(name):
    s1 = first_cap_re.sub(r'\1_\2', name)
    return all_cap_re.sub(r'\1_\2', s1).lower()


class BrowseView(generic.ListView):
    template_name = 'reports/reports.html'
    context_object_name = 'latest_report_list'

    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.selected_set = ''

    def get_queryset(self):
        """Return the last 11 published reports: excluding those published in the future."""
        reports = Report.objects.filter(pub_date__lte=timezone.now())
        self.selected_set = self.request.GET.get('set', '')
        if self.selected_set == 'cloud':
            reports = reports.filter(report_title__startswith='Cloud')
        elif self.selected_set == 'storage':
            reports = reports.filter(report_title__startswith='Storage')
        return reports.order_by('report_title')[:11]

    # we override this to add the selected set to the context (is there a better way?)
    def get_context_data(self, **kwargs):
        result = super().get_context_data(**kwargs)
        result['set'] = self.selected_set
        return result


class DetailView(generic.DetailView):
    model = Report
    template_name = 'reports/details.html'

    def get_queryset(self):
        """
        Excludes any reports that aren't published yet.
        """
        return Report.objects.filter(pub_date__lte=timezone.now())

    # we override this to add the debug flag to the context (is there a better way?)
    def get_context_data(self, **kwargs):
        result = super().get_context_data(**kwargs)
        result['debug'] = settings.DEBUG
        result['set'] = self.request.GET.get('set', '')
        return result


def about(request):
    values = sorted(request.META.items())
    return render(request, 'reports/about.html', {'details': values, 'debug': settings.DEBUG})


def search(request):
    q = request.GET.get('q', None)
    if not q:
        errors = ['Please enter a search term!']
        return render(request, 'reports/search_form.html', {'errors': errors})
    reports = Report.objects.filter(report_title__icontains=q)
    if reports.count() > 0:
        return render(request, 'reports/search_results.html', {'reports': reports, 'query': q})
    else:
        return render(request, 'reports/search_form.html', {'query': q})


def data(request, path):
    duration = request.GET.get('from', '')
    data_format = request.GET.get('format', 'json')
    filename = '%s%s.%s' % (path, duration, data_format)
    file_path = os.path.join(settings.BASE_DIR, 'reports', 'static', 'fake_data', filename)
    mime_type = 'application/json' if data_format == 'json' else 'text/csv'
    wrapper = FileWrapper(open(file_path))
    response = HttpResponse(wrapper, content_type=mime_type)
    # response['Content-Length'] = os.path.getsize(filename)
    # print("S %s" % os.path.getsize(filename))
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    return response


def manufactured(request, path):
    duration = request.GET.get('from', 'year')
    storage_type = request.GET.get('type', 'total')
    quota = factory.get(path, duration, storage_type)
    return JsonResponse(quota, safe=False, json_dumps_params={'indent': 2})


def xlsx(request, path):
    wb = Workbook(encoding='utf-8')
    ws = wb.worksheets[0]
    ws.title = 'range names'
    for col_idx in range(1, 20):
        col = get_column_letter(col_idx)
        for row in range(1, 40):
            ws.cell('%s%s' % (col, row)).value = '%s%s' % (col, row)
    ws = wb.create_sheet()
    ws.title = 'Pi'
    ws.cell('F5').value = 3.14
    return HttpResponse(save_virtual_workbook(wb), content_type='application/vnd.ms-excel')


def get_start_date(duration):
    if duration == 'oneMonth':
        return date.today() - relativedelta(months=1)
    elif duration == 'threeMonths':
        return date.today() - relativedelta(months=3)
    elif duration == 'sixMonths':
        return date.today() - relativedelta(months=6)
    else:
        # duration == 'year':
        return date.today() - relativedelta(months=12)


def actual(request, path):
    """
    :return: A response with the model data presented as CSV. The data is filtered till today by date range
    """
    duration = request.GET.get('from', 'year')
    desired_model = request.GET.get('model', 'unknown')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment;filename=%s.csv' % convert(desired_model)
    writer = csv.writer(response)
    # if not specified 'unknown' will not be found and an error raised
    model = django.apps.apps.get_model('reports', desired_model)
    field_names = [f.name for f in model._meta.fields]
    writer.writerow(field_names)
    for instance in model.objects.filter(date__range=((get_start_date(duration)), date.today())):
        writer.writerow([getattr(instance, f) for f in field_names])
    return response
